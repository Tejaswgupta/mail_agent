"""Detect airline carrier and normalize passenger manifests to a unified schema.

Supported carriers and formats:
  6E  IndiGo        – .xlsx / .csv, title row + header row
  IX  Air India Exp – .xls  / .csv, header row only
  TG  Thai Airways  – .xls  / .csv / .pdf, header row only

PDF extraction uses three strategies in priority order:
  1. Lines-based table extraction  (table borders drawn as lines)
  2. Text-based table extraction   (columns aligned, no lines)
  3. Word bounding-box reconstruction (arbitrary layout)
"""

import csv
import datetime
import io
import re
from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger

try:
    import xlrd  # type: ignore
    _XLRD_AVAILABLE = True
except ImportError:
    _XLRD_AVAILABLE = False

try:
    import pdfplumber  # type: ignore
    _PDFPLUMBER_AVAILABLE = True
except ImportError:
    _PDFPLUMBER_AVAILABLE = False

# Excel epoch for serial-date conversion (xlrd uses 1900 by default)
_XL_EPOCH = datetime.date(1899, 12, 30)


def _coerce_date(val: Any) -> str | None:
    """Convert an Excel serial float or various string formats to ISO date string."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, (int, float)):
        try:
            return (_XL_EPOCH + datetime.timedelta(days=int(val))).isoformat()
        except (OverflowError, ValueError):
            return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date().isoformat() if isinstance(val, datetime.datetime) else val.isoformat()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(s.split(" ")[0].split("T")[0], fmt).date().isoformat()
        except ValueError:
            continue
    return s or None


# ── Column maps (source header → unified field) ───────────────────────────────

_INDIGO_MAP: dict[str, str] = {
    "Flight Number":          "flight_number",
    "Flight Date":            "flight_date",
    "Origin Airport":         "origin",
    "Destination Airport":    "destination",
    "PNR":                    "pnr",
    "First Name":             "first_name",
    "Last Name":              "last_name",
    "Passenger Type":         "passenger_type",
    "Gender":                 "gender",
    "Nationality":            "nationality",
    "Passport No":            "passport_number",
    "Class of Travel(6E)":    "cabin_class",
    "Seat Number":            "seat_number",
    "No. of Bags":            "no_of_bags",
    "Baggage Weight (in kgs)":"baggage_weight",
    "Email Address":          "email",
    "Contact Number":         "phone",
    "Date of Booking":        "booking_date",
}

_AIX_MAP: dict[str, str] = {
    "FlightNumber":           "flight_number",
    "CarrierCode":            "_carrier_code",  # merged into flight_number post-mapping
    "DepartureDate":          "flight_date",
    "RecordLocator":          "pnr",
    "Title":                  "title",
    "BookingPerson1stName":   "first_name",
    "BookingPersonLastName":  "last_name",
    "PaxType":                "passenger_type",
    "Pax_DOB":                "date_of_birth",
    "passenger_nationality":  "nationality",
    "PPNo":                   "passport_number",
    "SeatNumber":             "seat_number",
    "ProductClassCode":       "cabin_class",
    "BaggageCount":           "no_of_bags",
    "weight":                 "baggage_weight",
    "Email":                  "email",
    "HomePhone":              "phone",
}

_TG_MAP: dict[str, str] = {
    "FLIGHT NUMBER":                              "flight_number",
    "OPERATION DATE":                             "flight_date",
    "EMBARK":                                     "origin",
    "DISEMBARK":                                  "destination",
    "PNR":                                        "pnr",
    "PAX LAST NAME":                              "last_name",
    "PAX FULL NAME":                              "full_name",
    "PHONE":                                      "phone",
    "EMAIL":                                      "email",
    "TRAVEL DOCUMENT NUMBER":                     "passport_number",
    "DATE OF BIRTH":                              "date_of_birth",
    "NATIONALITY":                                "nationality",
    "TICKET NO":                                  "ticket_number",
    "TICKET ISSUE DATE":                          "ticket_issue_date",
    "CABIN CLASS":                                "cabin_class",
    "SEAT NO":                                    "seat_number",
    "NO OF BAGS":                                 "no_of_bags",
    "BAGGAGE WEIGHT":                             "baggage_weight",
    "MODE OF PAYMENT (INCLUDE LAST 4 DIGIT OF CREDIT CARD USED)": "payment_mode",
}

_CARRIER_MAPS = {"6E": _INDIGO_MAP, "IX": _AIX_MAP, "TG": _TG_MAP}

# ── CSV aliases (header variants found in carrier CSV exports) ────────────────
# Each carrier may use slightly different column names in CSV vs Excel.
# We merge these into the canonical map at detection time.

_INDIGO_CSV_EXTRAS: dict[str, str] = {
    "Flt No":             "flight_number",
    "Flt Date":           "flight_date",
    "Dep":                "origin",
    "Arr":                "destination",
    "Class":              "cabin_class",
    "Seat":               "seat_number",
    "Bags":               "no_of_bags",
    "Bag Wt":             "baggage_weight",
    "Email":              "email",
    "Mobile":             "phone",
    "Booking Date":       "booking_date",
    "Passport":           "passport_number",
    "Type":               "passenger_type",
}

_AIX_CSV_EXTRAS: dict[str, str] = {
    "Flight Number":      "flight_number",
    "Carrier":            "_carrier_code",
    "Departure Date":     "flight_date",
    "PNR":                "pnr",
    "First Name":         "first_name",
    "Last Name":          "last_name",
    "PAX Type":           "passenger_type",
    "DOB":                "date_of_birth",
    "Nationality":        "nationality",
    "Passport":           "passport_number",
    "Seat":               "seat_number",
    "Class":              "cabin_class",
    "Bags":               "no_of_bags",
    "Baggage Weight":     "baggage_weight",
    "Phone":              "phone",
    "Origin":             "origin",
    "Destination":        "destination",
}

_TG_CSV_EXTRAS: dict[str, str] = {
    "Flight":             "flight_number",
    "Date":               "flight_date",
    "From":               "origin",
    "To":                 "destination",
    "Full Name":          "full_name",
    "Last Name":          "last_name",
    "Passport":           "passport_number",
    "Ticket Number":      "ticket_number",
    "Issue Date":         "ticket_issue_date",
    "Cabin":              "cabin_class",
    "Seat":               "seat_number",
    "Bags":               "no_of_bags",
    "Weight":             "baggage_weight",
    "Payment":            "payment_mode",
}

_CSV_EXTRA_MAPS: dict[str, dict[str, str]] = {
    "6E": _INDIGO_CSV_EXTRAS,
    "IX": _AIX_CSV_EXTRAS,
    "TG": _TG_CSV_EXTRAS,
}

# ── PDF keyword fingerprints (words that appear in the header region) ─────────
_PDF_FINGERPRINTS: dict[str, list[str]] = {
    "6E": ["INDIGO", "6E", "PREDEP", "PRE-DEPARTURE"],
    "IX": ["AIR INDIA EXPRESS", "AIR INDIA XPRESS", "IX", "RECORDLOCATOR"],
    "TG": ["THAI AIRWAYS", "THAI SMILE", "EMBARK", "DISEMBARK", "OPERATION DATE"],
}


# ── Carrier + manifest type detection ────────────────────────────────────────

def _detect_carrier(filename: str, headers: list[str], free_text: str = "") -> str | None:
    """Identify carrier from column headers, then free text, then filename."""
    header_set = set(h.strip() for h in headers)
    name_up = filename.upper()

    # Header-based detection (most reliable for Excel/CSV)
    if "RecordLocator" in header_set or "RECORDLOCATOR" in {h.upper() for h in header_set}:
        return "IX"
    if "PAX FULL NAME" in header_set or "EMBARK" in header_set:
        return "TG"
    if "Class of Travel(6E)" in header_set or "Origin Airport" in header_set:
        return "6E"
    # CSV extras detection via normalised header names
    hup = {h.upper() for h in header_set}
    if "FLIGHTNUMBER" in hup or "CARRIERCCODE" in hup:
        return "IX"

    # Free-text fingerprint (PDF / CSV with no canonical headers)
    if free_text:
        up = free_text.upper()
        for carrier, keywords in _PDF_FINGERPRINTS.items():
            if any(kw in up for kw in keywords):
                return carrier

    # Filename fallback
    if "IX" in name_up and ("MCT" in name_up or "BOM" in name_up or "MANIFEST" in name_up):
        return "IX"
    if name_up.startswith("TG") or "PRE-TG" in name_up or "-TG" in name_up:
        return "TG"
    if "6E" in name_up or name_up.startswith("PREDEP_6E"):
        return "6E"

    return None


def _detect_manifest_type(carrier: str, filename: str, title_text: str | None) -> str:
    name_up = filename.upper()
    if carrier == "6E" and title_text:
        return "pre_departure" if "PRE-DEPARTURE" in title_text.upper() else "post_departure"
    if "PREDEP" in name_up or name_up.startswith("PRE-"):
        return "pre_departure"
    return "post_departure"


# ── File readers ──────────────────────────────────────────────────────────────

def _read_xlsx(path: Path) -> tuple[str | None, list[str], list[dict]]:
    """Return (title_text_or_None, headers, raw_row_dicts)."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return None, [], []

    # IndiGo files have a single-cell title as row 0
    title_text: str | None = None
    header_start = 0
    first_nonempty = [v for v in all_rows[0] if v is not None]
    if len(first_nonempty) == 1 or (first_nonempty and str(first_nonempty[0]).startswith("Airline:")):
        title_text = str(first_nonempty[0]) if first_nonempty else None
        header_start = 1

    if header_start >= len(all_rows):
        return title_text, [], []

    headers = [str(h).strip() if h is not None else "" for h in all_rows[header_start]]

    data_rows = []
    for row in all_rows[header_start + 1:]:
        if all(c is None for c in row):
            continue
        data_rows.append(dict(zip(headers, row)))

    return title_text, headers, data_rows


def _read_xls(path: Path) -> tuple[str | None, list[str], list[dict]]:
    """Return (title_text_or_None, headers, raw_row_dicts). Requires xlrd<2."""
    if not _XLRD_AVAILABLE:
        raise ImportError("xlrd<2 is required to read .xls files — pip install 'xlrd<2'")

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)

    if ws.nrows == 0:
        return None, [], []

    title_text: str | None = None
    header_start = 0
    first_row = [ws.cell_value(0, c) for c in range(ws.ncols)]
    non_empty = [v for v in first_row if v != "" and v is not None]
    if len(non_empty) == 1 or (non_empty and str(non_empty[0]).startswith("Airline:")):
        title_text = str(non_empty[0]) if non_empty else None
        header_start = 1

    if header_start >= ws.nrows:
        return title_text, [], []

    headers = [str(ws.cell_value(header_start, c)).strip() for c in range(ws.ncols)]

    data_rows = []
    for row_idx in range(header_start + 1, ws.nrows):
        row = [ws.cell_value(row_idx, c) for c in range(ws.ncols)]
        if all(v == "" or v is None for v in row):
            continue
        data_rows.append(dict(zip(headers, row)))

    return title_text, headers, data_rows


# ── CSV reader ────────────────────────────────────────────────────────────────

def _read_csv(path: Path) -> tuple[str | None, list[str], list[dict]]:
    """Read a CSV manifest file.

    Tries common encodings and auto-detects the dialect (comma vs tab vs
    semicolon).  Returns (title_text_or_None, headers, raw_row_dicts).
    """
    raw_bytes = path.read_bytes()

    # Detect encoding: UTF-8-BOM → UTF-8 → Latin-1 (covers most airline exports)
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw_bytes.decode("latin-1", errors="replace")

    lines = text.splitlines()
    if not lines:
        return None, [], []

    # Sniff delimiter from the first non-empty line
    sample = "\n".join(lines[:20])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel  # fall back to standard comma

    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = list(reader)

    if not all_rows:
        return None, [], []

    # Detect single-cell title row (IndiGo style)
    title_text: str | None = None
    header_start = 0
    first_nonempty = [c for c in all_rows[0] if c.strip()]
    if len(first_nonempty) == 1 or (first_nonempty and first_nonempty[0].startswith("Airline:")):
        title_text = first_nonempty[0] if first_nonempty else None
        header_start = 1

    if header_start >= len(all_rows):
        return title_text, [], []

    headers = [h.strip() for h in all_rows[header_start]]

    data_rows: list[dict] = []
    for row in all_rows[header_start + 1:]:
        if not any(c.strip() for c in row):
            continue
        # Pad short rows so zip aligns correctly
        padded = row + [""] * max(0, len(headers) - len(row))
        data_rows.append(dict(zip(headers, padded)))

    return title_text, headers, data_rows


# ── PDF reader ────────────────────────────────────────────────────────────────

def _pdf_rows_from_table(page) -> list[list[str]]:
    """Strategy 1: use pdfplumber's table extractor (works when lines are drawn)."""
    tables = page.extract_tables()
    if not tables:
        return []
    # Use the largest table on the page
    best = max(tables, key=lambda t: len(t))
    return [[str(cell or "").strip() for cell in row] for row in best]


def _pdf_rows_from_text(page) -> list[list[str]]:
    """Strategy 2: cluster words into columns by x-coordinate proximity.

    Works on text-only PDFs where columns are visually aligned but no rule
    lines are drawn.  We collect all words, bin them by x0 into column buckets,
    then reconstruct rows from y-position.
    """
    words = page.extract_words(x_tolerance=4, y_tolerance=4)
    if not words:
        return []

    # Sort words by (top, x0)
    words_sorted = sorted(words, key=lambda w: (round(w["top"] / 4) * 4, w["x0"]))

    # Group by quantised y (row) → list of (x0, text)
    rows_by_y: dict[int, list[tuple[float, str]]] = {}
    for w in words_sorted:
        y_key = round(w["top"] / 4) * 4
        rows_by_y.setdefault(y_key, []).append((w["x0"], w["text"]))

    if not rows_by_y:
        return []

    # Determine column x-boundaries from the densest row (assumed to be header)
    all_xs = [x for row in rows_by_y.values() for x, _ in row]
    if not all_xs:
        return []

    # Cluster x values: sort, then gap > 20pt means new column
    sorted_xs = sorted(set(round(x / 5) * 5 for x in all_xs))
    col_starts: list[float] = [sorted_xs[0]]
    for x in sorted_xs[1:]:
        if x - col_starts[-1] > 20:
            col_starts.append(x)

    n_cols = len(col_starts)

    def _assign_col(x: float) -> int:
        for i in range(len(col_starts) - 1, -1, -1):
            if x >= col_starts[i] - 10:
                return i
        return 0

    result: list[list[str]] = []
    for y_key in sorted(rows_by_y):
        cells: list[str] = [""] * n_cols
        for x, text in rows_by_y[y_key]:
            col_idx = _assign_col(x)
            cells[col_idx] = (cells[col_idx] + " " + text).strip()
        if any(c for c in cells):
            result.append(cells)

    return result


def _pdf_matrix_to_dicts(matrix: list[list[str]]) -> tuple[list[str], list[dict]]:
    """Convert a list-of-lists matrix into (headers, list-of-row-dicts).

    Scans the first few rows to find the most likely header row (highest
    proportion of non-numeric, title-case or upper-case cells).
    """
    if not matrix:
        return [], []

    def _header_score(row: list[str]) -> float:
        non_empty = [c for c in row if c]
        if not non_empty:
            return 0.0
        score = sum(
            1 for c in non_empty
            if not re.match(r"^\d[\d\s./\-]*$", c)  # not pure numbers/dates
        )
        return score / len(non_empty)

    best_idx = max(range(min(5, len(matrix))), key=lambda i: _header_score(matrix[i]))
    headers = matrix[best_idx]
    data_rows: list[dict] = []
    for row in matrix[best_idx + 1:]:
        if not any(c for c in row):
            continue
        padded = row + [""] * max(0, len(headers) - len(row))
        data_rows.append(dict(zip(headers, padded)))
    return headers, data_rows


def _read_pdf(path: Path) -> tuple[str | None, list[str], list[dict], str]:
    """Parse a PDF manifest.  Returns (title_text, headers, rows, free_text).

    Tries three strategies per page in priority order:
      1. Table extraction  (pdfplumber line-based tables)
      2. Word-cluster text extraction
      3. Raw text fallback (returns empty rows but populated free_text for
         carrier detection)
    """
    if not _PDFPLUMBER_AVAILABLE:
        raise ImportError("pdfplumber is required for PDF parsing — pip install pdfplumber")

    all_matrix: list[list[str]] = []
    free_text_parts: list[str] = []
    title_text: str | None = None

    with pdfplumber.open(str(path)) as pdf:
        for page_num, page in enumerate(pdf.pages):
            # Always collect raw text for carrier detection / title extraction
            raw_text = page.extract_text() or ""
            free_text_parts.append(raw_text)

            if page_num == 0 and raw_text:
                # Use first non-empty line as potential title
                for line in raw_text.splitlines():
                    line = line.strip()
                    if line:
                        title_text = line
                        break

            # Strategy 1: line-based tables
            matrix = _pdf_rows_from_table(page)

            # Strategy 2: word-cluster if no table found
            if not matrix:
                matrix = _pdf_rows_from_text(page)

            if matrix:
                if not all_matrix:
                    all_matrix.extend(matrix)
                else:
                    # Subsequent pages: skip header row if it matches page-1 header
                    page_header = matrix[0] if matrix else []
                    start = 1 if page_header == all_matrix[0] else 0
                    all_matrix.extend(matrix[start:])

    free_text = "\n".join(free_text_parts)
    if not all_matrix:
        return title_text, [], [], free_text

    headers, data_rows = _pdf_matrix_to_dicts(all_matrix)
    return title_text, headers, data_rows, free_text


# ── Row normalization ─────────────────────────────────────────────────────────

_DATE_FIELDS = {"flight_date", "date_of_birth", "booking_date", "ticket_issue_date"}


def _build_col_map(carrier: str, format_hint: str) -> dict[str, str]:
    """Return the merged column map for a carrier, including format-specific aliases."""
    base = dict(_CARRIER_MAPS[carrier])
    if format_hint in ("csv", "pdf"):
        base.update(_CSV_EXTRA_MAPS.get(carrier, {}))
    return base


def _normalize_row(raw: dict, col_map: dict[str, str], carrier: str) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for src, dst in col_map.items():
        val = raw.get(src)
        if isinstance(val, str):
            val = val.strip() or None
        if val is None or val == "":
            continue
        if dst in _DATE_FIELDS:
            val = _coerce_date(val)
        if val is not None:
            row[dst] = val

    if carrier == "IX":
        carrier_code = row.pop("_carrier_code", "IX")
        fn = row.get("flight_number", "")
        if fn and not str(fn).startswith(str(carrier_code)):
            row["flight_number"] = f"{carrier_code}{int(fn) if isinstance(fn, float) else fn}"

    row["airline_code"] = carrier
    return row


# ── Public API ────────────────────────────────────────────────────────────────

def normalize(path: Path) -> tuple[str, str, list[dict[str, Any]]] | None:
    """Detect carrier and return normalized passenger rows.

    Supports .xlsx, .xls, .csv, and .pdf.

    Returns ``(airline_code, manifest_type, rows)`` where each row is a dict
    of unified fields plus a ``"_raw"`` key holding the original source values.
    Returns ``None`` if the file is not recognized as a passenger manifest.
    """
    suffix = path.suffix.lower()
    free_text = ""

    try:
        if suffix == ".xlsx":
            title_text, headers, raw_rows = _read_xlsx(path)
            format_hint = "xlsx"
        elif suffix == ".xls":
            title_text, headers, raw_rows = _read_xls(path)
            format_hint = "xls"
        elif suffix == ".csv":
            title_text, headers, raw_rows = _read_csv(path)
            format_hint = "csv"
        elif suffix == ".pdf":
            title_text, headers, raw_rows, free_text = _read_pdf(path)
            format_hint = "pdf"
        else:
            return None
    except Exception as exc:
        logger.warning(f"manifest_normalizer: cannot read {path.name}: {exc}")
        return None

    carrier = _detect_carrier(path.name, headers, free_text)
    if carrier is None:
        logger.debug(f"manifest_normalizer: unrecognized carrier in {path.name}")
        return None

    manifest_type = _detect_manifest_type(carrier, path.name, title_text)
    col_map = _build_col_map(carrier, format_hint)

    rows: list[dict[str, Any]] = []
    for raw in raw_rows:
        normalized = _normalize_row(raw, col_map, carrier)
        normalized["_raw"] = {k: str(v) for k, v in raw.items() if v not in (None, "")}
        rows.append(normalized)

    logger.info(
        f"manifest_normalizer: {path.name} [{format_hint}] "
        f"→ {carrier} / {manifest_type} / {len(rows)} row(s)"
    )
    return carrier, manifest_type, rows
