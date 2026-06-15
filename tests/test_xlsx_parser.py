"""Tests for xlsx_parser.py — builds real xlsx fixtures in tmp_path."""
from pathlib import Path
import pytest
import openpyxl
import xlsx_parser


def _make_xlsx(path: Path, sheets: dict) -> Path:
    """Create a minimal xlsx file. sheets = {name: [list_of_row_tuples]}"""
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(list(row))
    wb.save(str(path))
    return path


def test_basic_sheet_parsed(tmp_path):
    f = _make_xlsx(tmp_path / "test.xlsx", {
        "Invoices": [
            ("Invoice No", "Date", "Amount"),
            ("INV-001", "2024-01-01", 1000),
            ("INV-002", "2024-01-02", 2000),
        ]
    })
    result = xlsx_parser.parse(f)
    assert "Invoices" in result
    rows = result["Invoices"]
    assert len(rows) == 2
    assert rows[0]["Invoice No"] == "INV-001"
    assert rows[0]["Amount"] == 1000
    assert rows[1]["Invoice No"] == "INV-002"


def test_multiple_sheets(tmp_path):
    f = _make_xlsx(tmp_path / "multi.xlsx", {
        "Sheet1": [("A", "B"), (1, 2)],
        "Sheet2": [("X", "Y"), (3, 4)],
    })
    result = xlsx_parser.parse(f)
    assert set(result.keys()) == {"Sheet1", "Sheet2"}
    assert result["Sheet1"][0] == {"A": 1, "B": 2}
    assert result["Sheet2"][0] == {"X": 3, "Y": 4}


def test_blank_rows_skipped(tmp_path):
    f = _make_xlsx(tmp_path / "blanks.xlsx", {
        "Data": [
            ("Col1", "Col2"),
            ("a", "b"),
            (None, None),   # blank row — should be skipped
            ("c", "d"),
        ]
    })
    rows = xlsx_parser.parse(f)["Data"]
    assert len(rows) == 2
    assert rows[0]["Col1"] == "a"
    assert rows[1]["Col1"] == "c"


def test_empty_sheet_returns_no_entry(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    p = tmp_path / "empty.xlsx"
    wb.save(str(p))
    result = xlsx_parser.parse(p)
    assert result == {} or "Empty" not in result


def test_duplicate_headers_deduped(tmp_path):
    f = _make_xlsx(tmp_path / "dup.xlsx", {
        "Sheet1": [
            ("Name", "Name", "Value"),
            ("Alice", "A", 10),
        ]
    })
    rows = xlsx_parser.parse(f)["Sheet1"]
    keys = list(rows[0].keys())
    assert len(set(keys)) == len(keys)  # all unique
    assert "Name" in keys


def test_invalid_file_raises(tmp_path):
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"this is not an xlsx file")
    with pytest.raises(ValueError):
        xlsx_parser.parse(bad)
