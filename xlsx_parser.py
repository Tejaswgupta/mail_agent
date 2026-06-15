"""Parse xlsx attachments with a fixed schema into a list of row dicts.

Each sheet is parsed independently.  The first non-empty row is treated as the
header row.  Every subsequent row becomes a dict keyed by those headers.
Empty rows (all cells blank) are skipped.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from loguru import logger
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


def parse(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Return ``{sheet_name: [row_dict, ...]}`` for every sheet in the workbook.

    Raises ``ValueError`` if the file cannot be opened as a workbook.
    """
    result: dict[str, list[dict[str, Any]]] = {}

    if path.suffix.lower() == '.xls':
        import xlrd
        try:
            wb = xlrd.open_workbook(str(path))
            for sheet in wb.sheets():
                sheet_name = sheet.name
                rows = []
                for r in range(sheet.nrows):
                    rows.append(sheet.row_values(r))
                _parse_sheet_rows(sheet_name, rows, result)
        except Exception as exc:
            raise ValueError(f"Cannot open xls file {path.name}: {exc}") from exc
    else:
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                _parse_sheet_rows(sheet_name, rows, result)
            wb.close()
        except (InvalidFileException, Exception) as exc:
            raise ValueError(f"Cannot open xlsx file {path.name}: {exc}") from exc

    return result

def _parse_sheet_rows(sheet_name: str, rows: list, result: dict[str, list[dict[str, Any]]]) -> None:
    if not rows:
        logger.debug(f"Sheet '{sheet_name}' is empty — skipping")
        return

    # Find the first row that has at least one non-None/non-empty cell → headers
    header_idx = next(
        (i for i, r in enumerate(rows) if any(c is not None and str(c).strip() != "" for c in r)),
        None,
    )
    if header_idx is None:
        logger.debug(f"Sheet '{sheet_name}' has no non-empty rows — skipping")
        return

    raw_headers = rows[header_idx]
    # Deduplicate blank/None header names so dict keys are unique
    headers: list[str] = []
    seen: dict[str, int] = {}
    for h in raw_headers:
        key = str(h).strip() if (h is not None and str(h).strip() != "") else "column"
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        headers.append(key)

    data_rows: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # skip blank rows
        
        # pad row to match header length if needed
        row_list = list(row)
        if len(row_list) < len(headers):
            row_list.extend([None] * (len(headers) - len(row_list)))
        elif len(row_list) > len(headers):
            row_list = row_list[:len(headers)]
            
        data_rows.append(dict(zip(headers, row_list)))

    result[sheet_name] = data_rows
    logger.info(f"Parsed sheet '{sheet_name}': {len(data_rows)} row(s)")
